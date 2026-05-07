"""Testes do gerador de relatórios XLSX e PDF (Apêndice I)."""

import json
import pathlib

import openpyxl
import pytest

from src.report_builder import (
    _build_table_a,
    _build_table_b,
    _fmt_val,
    _g,
    _pivot_items,
    build_pdf,
    build_xlsx,
)
from src.siconfi_client import SiconfiResponse

FIXTURES_DIR = pathlib.Path(__file__).parent / "fixtures"


@pytest.fixture()
def arapiraca_response():
    with open(FIXTURES_DIR / "arapiraca_2025.json", encoding="utf-8") as f:
        data = json.load(f)
    items = data["items"]
    return SiconfiResponse(
        items=items,
        instituicao=items[0]["instituicao"],
        exercicio=2025,
        periodo=6,
        url_chamada="https://apidatalake.tesouro.gov.br/ords/siconfi/tt/rreo?...",
        metadados={"count": data["count"]},
        data_status="2026-01-30T11:34:12Z",
    )


# ── _pivot_items ──────────────────────────────────────────────────────────────

class TestPivotItems:
    def test_estrutura_basica(self):
        items = [
            {"cod_conta": "ReceitasCorrentes", "coluna": "PREVISÃO INICIAL", "valor": 100.0},
            {"cod_conta": "ReceitasCorrentes", "coluna": "Até o Bimestre (c)", "valor": 90.0},
            {"cod_conta": "DespesasCorrentes", "coluna": "DOTAÇÃO INICIAL (d)", "valor": 50.0},
        ]
        pivot = _pivot_items(items)
        assert pivot["ReceitasCorrentes"]["PREVISÃO INICIAL"] == 100.0
        assert pivot["ReceitasCorrentes"]["Até o Bimestre (c)"] == 90.0
        assert pivot["DespesasCorrentes"]["DOTAÇÃO INICIAL (d)"] == 50.0

    def test_valor_none_vira_zero(self):
        items = [{"cod_conta": "X", "coluna": "Y", "valor": None}]
        pivot = _pivot_items(items)
        assert pivot["X"]["Y"] == 0.0

    def test_ignora_item_sem_cod_conta(self):
        items = [{"cod_conta": "", "coluna": "Y", "valor": 1.0}]
        pivot = _pivot_items(items)
        assert pivot == {}

    def test_ignora_item_sem_coluna(self):
        items = [{"cod_conta": "X", "coluna": "", "valor": 1.0}]
        pivot = _pivot_items(items)
        assert pivot == {}

    def test_fixture_arapiraca_tem_contas_esperadas(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        assert "ReceitasCorrentes" in pivot
        assert "DespesasCorrentes" in pivot
        assert "DespesasDeCapital" in pivot


# ── _g ────────────────────────────────────────────────────────────────────────

class TestGetter:
    def test_retorna_valor_existente(self):
        pivot = {"ReceitasCorrentes": {"PREVISÃO INICIAL": 500.0}}
        assert _g(pivot, "ReceitasCorrentes", "PREVISÃO INICIAL") == 500.0

    def test_retorna_zero_quando_conta_ausente(self):
        assert _g({}, "Inexistente", "PREVISÃO INICIAL") == 0.0

    def test_retorna_zero_quando_coluna_ausente(self):
        pivot = {"X": {"A": 1.0}}
        assert _g(pivot, "X", "B") == 0.0

    def test_retorna_zero_quando_conta_none(self):
        assert _g({}, None, "PREVISÃO INICIAL") == 0.0

    def test_retorna_valor_fixture(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        val = _g(pivot, "ReceitasCorrentes", "PREVISÃO INICIAL")
        assert val == pytest.approx(1030471541, rel=1e-6)


# ── _fmt_val ──────────────────────────────────────────────────────────────────

class TestFmtVal:
    def test_formata_milhoes_com_separadores_brasileiros(self):
        assert _fmt_val(1000000.0) == "1.000.000,00"

    def test_formata_decimal_com_centavos(self):
        assert _fmt_val(1234.56) == "1.234,56"

    def test_zero_vira_zero_formatado(self):
        assert _fmt_val(0.0) == "0,00"

    def test_none_vira_zero_formatado(self):
        assert _fmt_val(None) == "0,00"

    def test_zero_percentual(self):
        assert _fmt_val(0.0, is_pct=True) == "0,00%"

    def test_formato_percentual(self):
        assert _fmt_val(0.1234, is_pct=True) == "12.34%"

    def test_negativo(self):
        assert _fmt_val(-500.0) == "-500,00"


# ── _build_table_a ────────────────────────────────────────────────────────────

class TestBuildTableA:
    def test_primeira_linha_e_cabecalho_com_sete_colunas(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, *_ = _build_table_a(pivot)
        assert table[0][0] == "Receitas"
        assert len(table[0]) == 7

    def test_tem_linha_receitas_correntes(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, *_ = _build_table_a(pivot)
        labels = [row[0] for row in table[1:]]
        assert "Receitas Correntes (I)" in labels

    def test_tem_linha_total_receitas(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, *_ = _build_table_a(pivot)
        labels = [row[0] for row in table[1:]]
        assert "TOTAL DAS RECEITAS (VIII) = (III+VI)" in labels

    def test_valores_receitas_correntes(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, *_ = _build_table_a(pivot)
        rc_row = next(r for r in table[1:] if r[0] == "Receitas Correntes (I)")
        assert rc_row[1] == _fmt_val(1030471541)      # PI
        assert rc_row[2] == _fmt_val(1218263018.8)    # PA
        assert rc_row[3] == _fmt_val(1267662215.61)   # AR

    def test_nao_tem_linha_receitas_orcamentarias(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, *_ = _build_table_a(pivot)
        labels = [row[0] for row in table[1:]]
        assert "Receitas Orçamentárias" not in labels

    def test_cada_linha_tem_sete_colunas(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, *_ = _build_table_a(pivot)
        for i, row in enumerate(table):
            assert len(row) == 7, f"Linha {i} tem {len(row)} colunas, esperado 7"


# ── _build_table_b ────────────────────────────────────────────────────────────

class TestBuildTableB:
    def test_primeira_linha_e_cabecalho_com_sete_colunas(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, _ = _build_table_b(pivot)
        assert table[0][0] == "Despesas"
        assert len(table[0]) == 7

    def test_tem_linha_despesas_correntes(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, _ = _build_table_b(pivot)
        labels = [row[0] for row in table[1:]]
        assert "Despesas Correntes (I)" in labels

    def test_tem_linha_total_despesas(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, _ = _build_table_b(pivot)
        labels = [row[0] for row in table[1:]]
        assert "TOTAL DAS DESPESAS (VIII) = (III+VII)" in labels

    def test_valores_despesas_correntes(self, arapiraca_response):
        pivot = _pivot_items(arapiraca_response.items)
        table, _ = _build_table_b(pivot)
        dc_row = next(r for r in table[1:] if r[0] == "Despesas Correntes (I)")
        assert dc_row[1] == _fmt_val(954927141)        # DI
        assert dc_row[2] == _fmt_val(1147906628.29)    # DA
        assert dc_row[3] == _fmt_val(1114371621.84)    # EM


# ── build_xlsx ────────────────────────────────────────────────────────────────

class TestBuildXlsx:
    def test_retorna_bytes_nao_vazios(self, arapiraca_response):
        data = build_xlsx(arapiraca_response)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_arquivo_e_xlsx_valido(self, arapiraca_response):
        import io as _io
        data = build_xlsx(arapiraca_response)
        wb = openpyxl.load_workbook(_io.BytesIO(data))
        assert wb.active is not None

    def test_label_receitas_correntes_na_celula_correta(self, arapiraca_response):
        import io as _io
        ws = openpyxl.load_workbook(_io.BytesIO(build_xlsx(arapiraca_response))).active
        # _fill_bloco_a: row=8 → "Receitas Correntes (I)", col=2
        assert ws.cell(row=8, column=2).value == "Receitas Correntes (I)"

    def test_receitas_correntes_pi_na_celula_correta(self, arapiraca_response):
        import io as _io
        ws = openpyxl.load_workbook(_io.BytesIO(build_xlsx(arapiraca_response))).active
        # row=8, col=3 → ReceitasCorrentes[PREVISÃO INICIAL] = 1030471541
        assert ws.cell(row=8, column=3).value == pytest.approx(1030471541, rel=1e-6)

    def test_despesas_correntes_di_na_celula_correta(self, arapiraca_response):
        import io as _io
        ws = openpyxl.load_workbook(_io.BytesIO(build_xlsx(arapiraca_response))).active
        # row=44, col=3 → DespesasCorrentes[DOTAÇÃO INICIAL (d)] = 954927141
        assert ws.cell(row=44, column=3).value == pytest.approx(954927141, rel=1e-6)

    def test_label_despesas_correntes_na_celula_correta(self, arapiraca_response):
        import io as _io
        ws = openpyxl.load_workbook(_io.BytesIO(build_xlsx(arapiraca_response))).active
        assert ws.cell(row=44, column=2).value == "Despesas Correntes (I)"

    def test_template_alternativo_aceito(self, arapiraca_response):
        data = build_xlsx(arapiraca_response, template_path=None)
        assert isinstance(data, bytes) and len(data) > 0


# ── build_pdf ─────────────────────────────────────────────────────────────────

class TestBuildPdf:
    def test_retorna_bytes_nao_vazios(self, arapiraca_response):
        data = build_pdf(arapiraca_response)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_arquivo_nao_vazio(self, arapiraca_response):
        data = build_pdf(arapiraca_response)
        assert len(data) > 1024

    def test_arquivo_comeca_com_assinatura_pdf(self, arapiraca_response):
        data = build_pdf(arapiraca_response)
        assert data[:4] == b"%PDF"

    def test_pdf_gerado_sem_data_status(self, arapiraca_response):
        arapiraca_response.data_status = None
        data = build_pdf(arapiraca_response)
        assert data[:4] == b"%PDF"

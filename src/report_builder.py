"""
Gerador de relatórios XLSX e PDF — Apêndice I (RREO Anexo 1, 6º bimestre).

Estratégia XLSX: carrega o template vazio, que contém apenas referências a arquivo
externo (=IF('[1]ORCAMENTO_A PREENCHER'!...)), substitui todas as células por
valores estáticos calculados a partir da API SICONFI e salva em output_path.
"""

import pathlib
from collections import defaultdict
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import numbers as xl_numbers
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .siconfi_client import SiconfiResponse

_TEMPLATE_PATH = (
    pathlib.Path(__file__).parent.parent / "templates" / "Modelo_Relatorio_Apendice_1.xlsx"
)

# ── Colunas da API ────────────────────────────────────────────────────────────
_PI = "PREVISÃO INICIAL"
_PA = "PREVISÃO ATUALIZADA (a)"
_AR = "Até o Bimestre (c)"
_DI = "DOTAÇÃO INICIAL (d)"
_DA = "DOTAÇÃO ATUALIZADA (e)"
_EM = "DESPESAS EMPENHADAS ATÉ O BIMESTRE (f)"
_LI = "DESPESAS LIQUIDADAS ATÉ O BIMESTRE (h)"
_PG = "DESPESAS PAGAS ATÉ O BIMESTRE (j)"
_RK = "INSCRITAS EM RESTOS A PAGAR NÃO PROCESSADOS (k)"

# ── Mapeamento: linha do template → (rótulo, cod_conta)  ─────────────────────
# Bloco A: colunas C=PI, D=PA, E=AR, F=diff(E-D), G=AV%, H=AH%
_BLOCO_A: list[tuple[int, str, str | None]] = [
    (6,  "Receitas",                                                       "__HEADER"),
    (7,  "Receitas Orçamentárias",                                         "ReceitasExcetoIntraOrcamentarias"),
    (8,  "Receitas Correntes (I)",                                         "ReceitasCorrentes"),
    (9,  "Impostos, Taxas e Contribuições de Melhoria",                    "ReceitaTributaria"),
    (10, "Contribuições",                                                   "ReceitaDeContribuicoes"),
    (11, "Receita Patrimonial",                                             "ReceitaPatrimonial"),
    (12, "Receita Agropecuária",                                            None),
    (13, "Receita Industrial",                                              None),
    (14, "Receita de Serviços",                                             "ReceitaDeServicos"),
    (15, "Transferências Correntes",                                        "TransferenciasCorrentes"),
    (16, "       Transf. da União e de suas Entidades",                    "TransferenciasCorrentesDaUniaoEDeSuasEntidades"),
    (17, "       Transf. dos Estados e do DF e de suas Entidades",         "TransferenciasCorrentesDosEstadosEDoDistritoFederalEDeSuasEntidades"),
    (18, "       Transf. de Outras Instituições Públicas",                 "TransferenciasCorrentesDeOutrasInstituicoesPublicas"),
    (19, "       Demais Transferências Correntes",                         "OutrasTransferenciasCorrentes"),
    (20, "Outras Receitas Correntes",                                       "OutrasReceitasCorrentes"),
    (21, "Receitas Intraorçamentárias Correntes (II)",                     "ReceitasCorrentesIntra"),
    (22, "Subtotal das Receitas Correntes (III) = (I+II)",                 "__SUM_R22"),
    (23, "Receitas de Capital (IV)",                                        "ReceitasDeCapital"),
    (24, "Operações de Crédito",                                            "ReceitasDeOperacoesDeCredito"),
    (25, "Alienação de Bens",                                               None),
    (26, "Amortização de Empréstimos",                                      None),
    (27, "Transferências de Capital",                                        "TransferenciasDeCapital"),
    (28, "       Transf. da União e de suas Entidades",                    "TransferenciasDeCapitalDaUniaoEDeSuasEntidades"),
    (29, "       Transf. dos Estados e de suas Entidades",                 None),
    (30, "       Transf. dos Municípios e de suas Entidades",              None),
    (31, "       Transf. de Instituições Privadas",                        None),
    (32, "       Transf. de Outras Instituições Públicas",                 None),
    (33, "       Transf. do Exterior",                                      None),
    (34, "       Demais Transferências de Capital",                         None),
    (35, "Outras Receitas de Capital",                                       None),
    (36, "Receitas Intraorçamentárias de Capital (V)",                     None),
    (37, "Subtotal das Receitas de Capital (VI) = (IV+V)",                 "__SUM_R37"),
    (38, "TOTAL DAS RECEITAS (VIII) = (III+VI)",                           "__SUM_R38"),
]

# Bloco B: colunas C=DI, D=DA, E=EM, F=LI, G=PG, H=RK
_BLOCO_B: list[tuple[int, str, str | None]] = [
    (43, "Despesas",                                                        "__HEADER"),
    (44, "Despesas Correntes (I)",                                          "DespesasCorrentes"),
    (45, "Pessoal e Encargos Sociais",                                      "PessoalEEncargosSociais"),
    (46, "Juros e Encargos da Dívida",                                      "JurosEEncargosDaDivida"),
    (47, "Outras Despesas Correntes",                                        "OutrasDespesasCorrentes"),
    (48, "Despesas Intraorçamentárias Correntes (II)",                      "DespesasCorrentesIntra"),
    (49, "Subtotal da Despesa Corrente (III) = (I+II)",                     "__SUM_B49"),
    (50, "Despesas de Capital (IV)",                                         "DespesasDeCapital"),
    (51, "Investimentos",                                                     "Investimentos"),
    (52, "Inversões Financeiras",                                             None),
    (53, "Amortização da Dívida",                                            "AmortizacaoDaDivida"),
    (54, "Reserva de Contingência (V)",                                      "ReservaDeContingencia"),
    (55, "Despesas Intraorçamentárias de Capital (VI)",                     "DespesasDeCapitalIntra"),
    (56, "Subtotal da Despesa de Capital (VII) = (IV+V+VI)",               "__SUM_B56"),
    (57, "TOTAL DAS DESPESAS (VIII) = (III+VII)",                           "__SUM_B57"),
    (58, "",                                                                  None),
    (59, "",                                                                  None),
]

_HEADER_A = ["Receitas", "Previsão Inicial (a)", "Previsão Atualizada (b)",
             "Receita Arrecadada (c)", "Diferença (c - b)", "A.V.%", "A.H.% (c/b)"]
_HEADER_B = ["Despesas", "Dotação Inicial", "Dotação Atualizada",
             "Empenhado", "Liquidado", "Pago", "RPNP"]

_CURRENCY_FMT = '#,##0.00'
_PCT_FMT      = '0.00%'

_BOLD_ROWS_A = {8, 21, 22, 23, 36, 37, 38}
_BOLD_ROWS_B = {44, 48, 49, 50, 55, 56, 57}


# ── Pivot ─────────────────────────────────────────────────────────────────────

def _pivot_items(items: list[dict[str, Any]]) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = defaultdict(dict)
    for item in items:
        cod = item.get("cod_conta") or ""
        col = item.get("coluna") or ""
        val = item.get("valor")
        if cod and col:
            result[cod][col] = float(val) if val is not None else 0.0
    return dict(result)


def _g(pivot: dict, conta: str | None, coluna: str) -> float:
    if not conta:
        return 0.0
    v = pivot.get(conta, {}).get(coluna)
    return float(v) if v else 0.0


# ── Valores por linha ─────────────────────────────────────────────────────────

def _receita_vals(pivot: dict, conta: str | None, total_ar: float) -> tuple[float, ...]:
    """Retorna (PI, PA, AR, diff, av_pct, ah_pct) para uma linha de receita."""
    pi = _g(pivot, conta, _PI)
    pa = _g(pivot, conta, _PA)
    ar = _g(pivot, conta, _AR)
    diff = ar - pa
    av   = ar / total_ar if total_ar else 0.0
    ah   = ar / pa       if pa       else 0.0
    return pi, pa, ar, diff, av, ah


def _despesa_vals(pivot: dict, conta: str | None) -> tuple[float, ...]:
    """Retorna (DI, DA, EM, LI, PG, RK) para uma linha de despesa."""
    di = _g(pivot, conta, _DI)
    da = _g(pivot, conta, _DA)
    em = _g(pivot, conta, _EM)
    li = _g(pivot, conta, _LI)
    pg = _g(pivot, conta, _PG)
    rk = _g(pivot, conta, _RK)
    return di, da, em, li, pg, rk


# ── XLSX ──────────────────────────────────────────────────────────────────────

def _set(ws, row: int, col: int, value: Any, fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if fmt:
        cell.number_format = fmt


def _fill_bloco_a(ws, pivot: dict) -> dict[str, list[float]]:
    """
    Preenche as linhas 6-38 do Bloco A no worksheet.
    Retorna cache {row_key: [pi, pa, ar]} para uso nas linhas calculadas.
    """
    total_ar = (
        _g(pivot, "ReceitasCorrentes", _AR)
        + _g(pivot, "ReceitasCorrentesIntra", _AR)
        + _g(pivot, "ReceitasDeCapital", _AR)
    )

    cache: dict[int, tuple[float, ...]] = {}

    for row, label, conta in _BLOCO_A:
        if conta == "__HEADER":
            for col, txt in enumerate(_HEADER_A, start=2):
                _set(ws, row, col, txt)
            continue

        if conta == "__SUM_R22":
            pi = cache[8][0] + cache[21][0]
            pa = cache[8][1] + cache[21][1]
            ar = cache[8][2] + cache[21][2]
            diff, av, ah = ar - pa, ar / total_ar if total_ar else 0.0, ar / pa if pa else 0.0
            vals = pi, pa, ar, diff, av, ah
        elif conta == "__SUM_R37":
            pi = cache[23][0] + cache[36][0]
            pa = cache[23][1] + cache[36][1]
            ar = cache[23][2] + cache[36][2]
            diff, av, ah = ar - pa, ar / total_ar if total_ar else 0.0, ar / pa if pa else 0.0
            vals = pi, pa, ar, diff, av, ah
        elif conta == "__SUM_R38":
            pi = cache[22][0] + cache[37][0]
            pa = cache[22][1] + cache[37][1]
            ar = cache[22][2] + cache[37][2]
            diff, av, ah = ar - pa, 1.0 if total_ar else 0.0, ar / pa if pa else 0.0
            vals = pi, pa, ar, diff, av, ah
        else:
            vals = _receita_vals(pivot, conta, total_ar)

        cache[row] = vals
        _set(ws, row, 2, label)
        _set(ws, row, 3, vals[0], _CURRENCY_FMT)
        _set(ws, row, 4, vals[1], _CURRENCY_FMT)
        _set(ws, row, 5, vals[2], _CURRENCY_FMT)
        _set(ws, row, 6, vals[3], _CURRENCY_FMT)
        _set(ws, row, 7, vals[4], _PCT_FMT)
        _set(ws, row, 8, vals[5], _PCT_FMT)

    return cache


def _fill_bloco_b(ws, pivot: dict) -> dict[int, tuple[float, ...]]:
    cache: dict[int, tuple[float, ...]] = {}

    for row, label, conta in _BLOCO_B:
        if conta == "__HEADER":
            for col, txt in enumerate(_HEADER_B, start=2):
                _set(ws, row, col, txt)
            continue

        if conta == "__SUM_B49":
            vals = tuple(a + b for a, b in zip(cache[44], cache[48]))
        elif conta == "__SUM_B56":
            vals = tuple(a + b + c for a, b, c in zip(cache[50], cache[54], cache[55]))
        elif conta == "__SUM_B57":
            vals = tuple(a + b for a, b in zip(cache[49], cache[56]))
        else:
            vals = _despesa_vals(pivot, conta)

        cache[row] = vals
        _set(ws, row, 2, label)
        for col_idx, val in enumerate(vals, start=3):
            _set(ws, row, col_idx, val, _CURRENCY_FMT)

    return cache


def _fill_bloco_c(ws, pivot: dict, receita_cache: dict, despesa_cache: dict) -> None:
    # No empty template, Bloco C usa linhas 64-67 (refs externas: A58-A61)
    total_ar  = receita_cache.get(38, (0,) * 6)[2]  # AR do TOTAL DAS RECEITAS
    total_em  = despesa_cache.get(57, (0,) * 6)[2]  # EM do TOTAL DAS DESPESAS
    resultado = total_ar - total_em

    bloco_c = [
        (64, ""),
        (65, "RECEITAS REALIZADAS"),
        (66, "DESPESAS EXECUTADAS (Empenhadas)"),
        (67, "RESULTADO ORÇAMENTÁRIO"),
    ]
    valores_c = [None, total_ar, total_em, resultado]

    for (row, label), val in zip(bloco_c, valores_c):
        _set(ws, row, 2, label)
        if val is not None:
            _set(ws, row, 3, val, _CURRENCY_FMT)


def build_xlsx(
    response: SiconfiResponse,
    output_path: str | pathlib.Path,
    template_path: str | pathlib.Path | None = None,
) -> pathlib.Path:
    """
    Preenche o template XLSX com os dados da SiconfiResponse e salva em output_path.

    Args:
        response: objeto retornado por fetch_rreo_anexo1().
        output_path: caminho de destino do XLSX gerado.
        template_path: caminho do template vazio (padrão: templates/Modelo_Relatorio_Apendice_1.xlsx).

    Returns:
        pathlib.Path do arquivo gerado.
    """
    tpl = pathlib.Path(template_path or _TEMPLATE_PATH)
    wb  = openpyxl.load_workbook(tpl)
    ws  = wb.active

    pivot = _pivot_items(response.items)

    titulo = f"{response.instituicao}  –  Exercício de {response.exercicio}"
    for row in (4, 41, 62):
        cell = ws.cell(row=row, column=2)
        if cell.value and ("XXXXXXX" in str(cell.value) or "XXXX" in str(cell.value)):
            cell.value = titulo

    receita_cache = _fill_bloco_a(ws, pivot)
    despesa_cache = _fill_bloco_b(ws, pivot)
    _fill_bloco_c(ws, pivot, receita_cache, despesa_cache)

    out = pathlib.Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ── PDF ───────────────────────────────────────────────────────────────────────

def _fmt_data_status(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.strptime(iso[:19], "%Y-%m-%dT%H:%M:%S")
        return dt.strftime("%d/%m/%Y às %H:%M:%S")
    except ValueError:
        return iso


def _fmt_val(v: float | None, is_pct: bool = False) -> str:
    if v is None or v == 0.0:
        return "0,00%" if is_pct else "0,00"
    if is_pct:
        return f"{v * 100:.2f}%"
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _build_table_a(pivot: dict) -> tuple[list[list], list[int]]:
    total_ar = (
        _g(pivot, "ReceitasCorrentes", _AR)
        + _g(pivot, "ReceitasCorrentesIntra", _AR)
        + _g(pivot, "ReceitasDeCapital", _AR)
    )
    rows: list[list] = [["Receitas", "Prev. Inicial (a)", "Prev. Atualizada (b)",
                          "Arrec. (c)", "Diferença (c-b)", "A.V.%", "A.H.%"]]
    bold_rows: list[int] = []
    neg_diff_rows: list[int] = []

    sums: dict[int, tuple[float, ...]] = {}
    for row_num, label, conta in _BLOCO_A:
        if conta == "__HEADER":
            continue
        if conta == "__SUM_R22":
            pi, pa, ar = (sums[8][i] + sums[21][i] for i in range(3))
        elif conta == "__SUM_R37":
            pi, pa, ar = (sums[23][i] + sums[36][i] for i in range(3))
        elif conta == "__SUM_R38":
            pi, pa, ar = (sums[22][i] + sums[37][i] for i in range(3))
        else:
            pi = _g(pivot, conta, _PI)
            pa = _g(pivot, conta, _PA)
            ar = _g(pivot, conta, _AR)

        sums[row_num] = (pi, pa, ar)

        if row_num == 7:  # "Receitas Orçamentárias" removida do PDF
            continue

        diff = ar - pa
        av   = ar / total_ar if total_ar else 0.0
        ah   = ar / pa       if pa       else 0.0

        if row_num in _BOLD_ROWS_A:
            bold_rows.append(len(rows))
        if diff < 0:
            neg_diff_rows.append(len(rows))

        rows.append([label, _fmt_val(pi), _fmt_val(pa), _fmt_val(ar),
                     _fmt_val(diff), _fmt_val(av, True), _fmt_val(ah, True)])
    return rows, bold_rows, neg_diff_rows


def _build_table_b(pivot: dict) -> tuple[list[list], list[int]]:
    rows: list[list] = [["Despesas", "Dot. Inicial", "Dot. Atualizada",
                          "Empenhado", "Liquidado", "Pago", "RPNP"]]
    bold_rows: list[int] = []
    sums: dict[int, tuple[float, ...]] = {}
    for row_num, label, conta in _BLOCO_B:
        if conta == "__HEADER" or label == "":
            continue
        if conta == "__SUM_B49":
            vals = tuple(sums[44][i] + sums[48][i] for i in range(6))
        elif conta == "__SUM_B56":
            vals = tuple(sums[50][i] + sums[54][i] + sums[55][i] for i in range(6))
        elif conta == "__SUM_B57":
            vals = tuple(sums[49][i] + sums[56][i] for i in range(6))
        else:
            vals = _despesa_vals(pivot, conta)
        sums[row_num] = vals

        if row_num in _BOLD_ROWS_B:
            bold_rows.append(len(rows))

        rows.append([label] + [_fmt_val(v) for v in vals])
    return rows, bold_rows


def _make_table_style(bold_rows: list[int], neg_diff_rows: list[int] | None = None) -> TableStyle:
    cmds: list = [
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#2E5E8E")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Times-Bold"),
        ("FONTNAME",      (0, 1), (-1, -1), "Times-Roman"),
        ("FONTSIZE",      (0, 0), (-1, -1), 10),
        ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN",         (0, 0), (0, -1),  "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F7")]),
        ("GRID",          (0, 0), (-1, -1), 0.25, colors.HexColor("#BBBBBB")),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
    ]
    for r in bold_rows:
        cmds.append(("FONTNAME", (0, r), (-1, r), "Times-Bold"))
    for r in (neg_diff_rows or []):
        cmds.append(("TEXTCOLOR", (4, r), (4, r), colors.red))
    return TableStyle(cmds)


def build_pdf(
    response: SiconfiResponse,
    output_path: str | pathlib.Path,
) -> pathlib.Path:
    """
    Gera o Apêndice I em PDF usando reportlab.

    Args:
        response: objeto retornado por fetch_rreo_anexo1().
        output_path: caminho de destino do PDF gerado.

    Returns:
        pathlib.Path do arquivo gerado.
    """
    out = pathlib.Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    pivot = _pivot_items(response.items)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading2"],
                                 fontName="Times-Bold", fontSize=10, spaceAfter=4)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontName="Times-Roman", fontSize=10)

    story = []
    story.append(Paragraph("APÊNDICE I", title_style))
    story.append(Paragraph(
        f"{response.instituicao} – Exercício de {response.exercicio}",
        sub_style,
    ))
    if response.data_status:
        story.append(Paragraph(f"Data de envio dos dados ao SICONFI: {_fmt_data_status(response.data_status)}", sub_style))
    story.append(Spacer(1, 0.3 * cm))

    col_widths_a = [8.0 * cm] + [3.2 * cm] * 6
    col_widths_b = [8.0 * cm] + [3.2 * cm] * 6

    story.append(Paragraph("Bloco A – Demonstrativo Analítico da Execução da Receita", title_style))
    data_a, bold_a, neg_a = _build_table_a(pivot)
    tbl_a = Table(data_a, colWidths=col_widths_a, repeatRows=1)
    tbl_a.setStyle(_make_table_style(bold_a, neg_a))
    story.append(tbl_a)
    story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph("Bloco B – Demonstrativo Analítico da Execução da Despesa", title_style))
    data_b, bold_b = _build_table_b(pivot)
    tbl_b = Table(data_b, colWidths=col_widths_b, repeatRows=1)
    tbl_b.setStyle(_make_table_style(bold_b))
    story.append(tbl_b)
    story.append(Spacer(1, 0.4 * cm))

    total_ar = (
        _g(pivot, "ReceitasCorrentes", _AR)
        + _g(pivot, "ReceitasCorrentesIntra", _AR)
        + _g(pivot, "ReceitasDeCapital", _AR)
    )
    total_em = _g(pivot, "TotalDespesas", _EM)
    resultado = total_ar - total_em
    sinal = "SUPERÁVIT" if resultado >= 0 else "DÉFICIT"

    data_c = [
        ["Bloco C – Resultado Orçamentário", "Valor (R$)"],
        ["Receitas Realizadas",            _fmt_val(total_ar)],
        ["Despesas Executadas (Empenhadas)", _fmt_val(total_em)],
        [f"{sinal} Orçamentário",           _fmt_val(abs(resultado))],
    ]
    story.append(Paragraph("Bloco C – Resultado Orçamentário", title_style))
    tbl_c = Table(data_c, colWidths=[10 * cm, 5 * cm])
    tbl_c.setStyle(_make_table_style([]))
    story.append(tbl_c)

    doc = SimpleDocTemplate(
        str(out),
        pagesize=landscape(A4),
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )
    doc.build(story)
    return out
